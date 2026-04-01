"""
Automação para Assinatura de Lotes DOU — SIFAMA / ANTT
Desenvolvido para o projeto CCOBI - SERASA
"""

import tkinter as tk
from tkinter import ttk, messagebox
import threading
import time
import os
from datetime import datetime

import sys

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException, NoSuchElementException,
    ElementClickInterceptedException, WebDriverException,
    StaleElementReferenceException,
)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    OPENPYXL_DISPONIVEL = True
except ImportError:
    OPENPYXL_DISPONIVEL = False


def _formatar_data_dd_mm_yyyy(texto: str) -> str:
    """Formata apenas dígitos em dd/mm/yyyy. Ex: '18022026' -> '18/02/2026'."""
    digitos = "".join(c for c in texto if c.isdigit())[:8]
    if len(digitos) <= 2:
        return digitos
    if len(digitos) <= 4:
        return digitos[:2] + "/" + digitos[2:]
    return digitos[:2] + "/" + digitos[2:4] + "/" + digitos[4:]



# ─────────────────────────────────────────────────────────────
# Constantes — URLs e IDs dos elementos
# ─────────────────────────────────────────────────────────────

URL_LOGIN  = "https://appweb1.antt.gov.br/sca/Site/Login.aspx"
URL_LOTES  = "https://appweb1.antt.gov.br/spm/Site/PublicacaoDOU/ConsultarAssinarLotePublicacaoSimplesDOU.aspx"

_CP = "ContentPlaceHolderCorpo_" * 4  # prefixo comum a 4 níveis

# Login
ID_LOGIN_USUARIO = f"{_CP}TextBoxUsuario"
ID_LOGIN_SENHA   = f"{_CP}TextBoxSenha"
ID_LOGIN_BTN     = f"{_CP}ButtonOk"

# Filtros da tela principal
ID_DATA_INICIAL    = f"{_CP}txbDataInicial"
ID_DATA_FINAL      = f"{_CP}txbDataFinal"
ID_TIPO_PUBLICACAO = f"{_CP}ddlTipoPublicacao"
ID_FORMA_FISCALIZACAO   = f"{_CP}ddlFormaFiscalizacao"
ID_TIPO_FISCALIZACAO    = f"{_CP}tipoSubTipoFiscalizacao_ddlTipoFiscalizacao"
ID_SUBTIPO_FISCALIZACAO = f"{_CP}tipoSubTipoFiscalizacao_ddlSubTipoFiscalizacao"
ID_BTN_PESQUISAR   = f"{_CP}btnPesquisar"
ID_TABELA_LOTES    = f"{_CP}gdvLotePublicacao"
CSS_BTN_ASSINAR    = f"[id^='{_CP}gdvLotePublicacao_btnAssinarAuto_']"
ID_BTN_PROX_PAG    = f"{_CP}ucPaginador_ucPaginador_lbNextPage"

# Tela de assinatura (nova guia)
ID_SENHA_CERT    = "ContentPlaceHolderCorpo_" * 3 + "ucSenhaCertificadoDigital_txbSenhaCertificadoDigital"
ID_BTN_SALVAR    = "ContentPlaceHolderCorpo_" * 3 + "btnSalvar"
ID_MESSAGEBOX_OK = "MessageBox_ButtonOk"

TIPOS_PUBLICACAO = {
    "Notificação de Autuação":      "1",
    "Notificação de Multa":         "2",
    "Cancelamento":                 "4",
    "Notificação de Segunda Multa": "10",
    "Notificação de Penalidade":    "16",
    "Notificação Final de Multa":   "17",
}

# Forma de Fiscalização (valor enviado ao sistema -> texto no combo)
FORMAS_FISCALIZACAO = {
    "": "--Selecione--",
    "1": "Eletrônica",
    "2": "Manual",
    "3": "Remota",
}
# Ordem para exibição no combo (valor, texto)
LISTA_FORMAS = [("", "--Selecione--"), ("1", "Eletrônica"), ("2", "Manual"), ("3", "Remota")]

# Tipo Fiscalização (mesma lista para todas as formas)
TIPOS_FISCALIZACAO = {
    "": "--Selecione--",
    "2": "Excesso de Peso",
    "3": "Cargas",
    "4": "Passageiros",
    "5": "Cargas Internacional",
    "7": "Passageiros Internacional",
    "8": "Infraestrutura Rodoviária",
    "9": "Evasão de Pedágio",
}
LISTA_TIPOS_FISC = [("", "--Selecione--"), ("2", "Excesso de Peso"), ("3", "Cargas"), ("4", "Passageiros"),
    ("5", "Cargas Internacional"), ("7", "Passageiros Internacional"), ("8", "Infraestrutura Rodoviária"), ("9", "Evasão de Pedágio")]

# Subtipo por Tipo (chave = value do Tipo; valor = lista de (value, texto))
SUBTIPOS_POR_TIPO = {
    "2": [("", "--Selecione--"), ("5", "Excesso de Peso"), ("7", "CMT - Capacidade Máxima de Tração"), ("16", "Evasão de Balança")],
    "3": [("", "--Selecione--"), ("8", "RNTRC - Registro Nacional de Transportadores Rodoviários de Cargas"), ("9", "PEF - Pagamento Eletrônico de Frete"), ("10", "Vale Pedágio"), ("17", "Produtos Perigosos"), ("24", "Piso Mínimo de Frete")],
    "4": [("", "--Selecione--"), ("11", "Longa Distância"), ("12", "Semiurbano"), ("13", "Fretamento"), ("14", "Não Autorizado"), ("19", "Passageiro Econômico Financeiro"), ("20", "Fretamento Contínuo"), ("23", "Ferroviário de Passageiros")],
    "5": [("", "--Selecione--"), ("21", "Cargas Internacional"), ("22", "Produtos Perigosos Internacional")],
    "7": [("", "--Selecione--"), ("25", "Longa Distância"), ("26", "Semiurbano"), ("27", "Fretamento"), ("28", "Não Autorizado"), ("29", "Fretamento Contínuo")],
    "8": [("", "--Selecione--"), ("30", "Infraestrutura Rodoviária")],
    "9": [("", "--Selecione--"), ("31", "Evasão de Pedágio")],
}

# Delays padrão por etapa (em segundos) — multiplicados pelos fatores da GUI
DELAYS = {
    "apos_clicar_assinar":      1.2,  # aguardar nova guia abrir
    "carregar_guia_assinatura": 2.0,  # aguardar tela de assinatura carregar completamente
    "apos_preencher_senha":     0.8,  # pausa antes de clicar Salvar
    "aguardar_messagebox":      600,  # timeout máximo (segundos) para o botão OK aparecer — 10 min
    "log_aguardando_ok_cada":   15,   # logar "ainda aguardando" a cada N segundos
    "apos_progresso_sumir":     1.5,  # pausa após a barra de progresso sumir, antes de clicar OK
    "apos_clicar_ok":           1.0,  # aguardar guia fechar após clicar OK
    "apos_voltar_aba":          1.5,  # estabilizar aba principal antes do próximo lote
    "entre_lotes":              1.5,  # pausa entre um lote e o próximo
    "carregar_proxima_pagina":  2.0,  # aguardar nova página da paginação carregar
    "apos_refresh":            4.0,   # (não usado — substituído por fechar/reabrir)
    "apos_fechar_navegador":   2.0,   # pausa após fechar navegador antes de reabrir (libera recursos)
}

# Chaves que recebem o fator_senha_cliques (senha + cliques) — permitem acelerar separadamente
DELAYS_SENHA_CLIQUES = frozenset({
    "apos_clicar_assinar", "carregar_guia_assinatura", "apos_preencher_senha",
    "apos_progresso_sumir", "apos_clicar_ok", "apos_voltar_aba", "entre_lotes",
})

# ─────────────────────────────────────────────────────────────
# O QUE É CONSIDERADO ERRO (apenas estas 3 situações):
#   1. Nova guia de assinatura não abriu em 60s após clicar Assinar
#   2. Barra Progress_UpdateProgress não sumiu em 600s após clicar Salvar
#      (ex.: senha errada, sistema fora do ar, travamento)
#   3. Exceção inesperada durante o processamento do lote
#
# NÃO É ERRO:
#   - Overlay wings_process_presentation_dashboard_bar visível (página carregando)
#   - Barra Progress_UpdateProgress visível (sistema processando) ← normal/esperado
#   - Guia não fechar automaticamente (fecha manualmente, continua normalmente)
#   - Delays / sleeps em andamento
# ─────────────────────────────────────────────────────────────


# ─────────────────────────────────────────────────────────────
# Logger
# ─────────────────────────────────────────────────────────────

class Logger:
    def __init__(self, callback=None, log_file: str = None):
        self.callback = callback
        self.log_file = log_file
        if log_file:
            pasta = os.path.dirname(log_file)
            if pasta:
                os.makedirs(pasta, exist_ok=True)
            with open(log_file, "w", encoding="utf-8") as f:
                f.write(f"Log de Automação — Assinatura de Lotes DOU\n")
                f.write(f"Iniciado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
                f.write("=" * 60 + "\n\n")

    def log(self, mensagem: str, tipo: str = "INFO"):
        ts = datetime.now().strftime("%H:%M:%S")
        linha = f"[{ts}] [{tipo}] {mensagem}"
        print(linha)
        if self.callback:
            self.callback(linha, tipo)
        if self.log_file:
            try:
                with open(self.log_file, "a", encoding="utf-8") as f:
                    f.write(linha + "\n")
            except Exception:
                pass


# ─────────────────────────────────────────────────────────────
# Automação — lógica de negócio
# ─────────────────────────────────────────────────────────────

class AutomacaoAssinaturaLotes:
    def __init__(
        self,
        logger: Logger,
        fator_delay: float = 1.0,
        fator_senha_cliques: float = 1.0,
    ):
        self.logger             = logger
        self.fator_delay        = fator_delay         # 1.0 = normal  |  2.0 = dobra todos os delays
        self.fator_senha_cliques = fator_senha_cliques  # 0.5 = senha/cliques mais rápido  |  1.0 = normal
        self.driver             = None
        self.wait               = None
        self.pausado            = False
        self.parar              = False
        self.resultados: list[dict] = []
        self.mensagem_final     = None  # Ex.: "Sessão expirada" — usado por _executar ao finalizar

    def _verificar_sessao_expirada(self) -> bool:
        """Retorna True se a página atual for a tela de login (sessão expirada)."""
        try:
            if not self.driver:
                return False
            url = self.driver.current_url or ""
            return "Login" in url or "login" in url.lower()
        except Exception:
            return False

    def _fechar_guia_assinatura_orphan(self):
        """Fecha guias extras (ex.: de assinatura deixada aberta após erro) e volta para a primeira."""
        try:
            if not self.driver or not self.driver.window_handles:
                return
            principal = self.driver.window_handles[0]
            for h in list(self.driver.window_handles):
                if h != principal:
                    try:
                        self.driver.switch_to.window(h)
                        self.driver.close()
                    except Exception:
                        pass
            self.driver.switch_to.window(self.driver.window_handles[0])
        except Exception:
            pass

    def _sleep(self, chave: str):
        """Dorme pelo tempo configurado para a etapa, respeitando pausa/parar."""
        segundos = DELAYS[chave] * self.fator_delay
        if chave in DELAYS_SENHA_CLIQUES:
            segundos *= self.fator_senha_cliques
        self.logger.log(f"   ⏱  aguardando {segundos:.1f}s ({chave})...", "INFO")
        fim = time.time() + segundos
        while time.time() < fim:
            if self.parar:
                return
            time.sleep(0.2)

    def _aguardar_pagina_pronta(self, timeout: int = 30, contexto: str = ""):
        """
        Aguarda os overlays de carregamento do sistema sumirem antes de interagir.

        Dois elementos cobrem a página durante o processamento:
          - Progress_UpdateProgress          → spinner "Processando..."
          - Progress_ModalProgress_backgroundElement → fundo opaco full-page (z-index 10000)

        Ambos precisam estar invisíveis para que cliques funcionem.

        NOTA: wings_process_presentation_dashboard_bar é rodapé FIXO permanente —
        NUNCA aguardar esse elemento.
        """
        prefixo = f"[{contexto}] " if contexto else ""

        # IDs que bloqueiam interações quando visíveis
        OVERLAYS = [
            "Progress_ModalProgress_backgroundElement",  # fundo opaco full-page
            "Progress_UpdateProgress",                   # spinner de processamento
        ]

        for id_overlay in OVERLAYS:
            try:
                el = self.driver.find_element(By.ID, id_overlay)
                try:
                    visivel = el.is_displayed()
                except StaleElementReferenceException:
                    visivel = False
                if visivel:
                    self.logger.log(
                        f"{prefixo}Overlay '{id_overlay}' visível — aguardando sumir...", "INFO"
                    )
                    WebDriverWait(self.driver, timeout, poll_frequency=0.3).until(
                        EC.invisibility_of_element_located((By.ID, id_overlay))
                    )
                    self.logger.log(f"{prefixo}Overlay '{id_overlay}' sumiu.", "INFO")
            except NoSuchElementException:
                pass  # não existe na página, ok
            except TimeoutException:
                self.logger.log(
                    f"{prefixo}Timeout {timeout}s aguardando '{id_overlay}' — continuando.", "WARNING"
                )

    def _esconder_rodape_fixo(self):
        """Esconde o rodapé fixo wings_process_presentation_dashboard_bar que intercepta cliques."""
        self.driver.execute_script(
            "var b=document.getElementById('wings_process_presentation_dashboard_bar');"
            "if(b){b.style.display='none';b.style.visibility='hidden';"
            "b.style.pointerEvents='none';b.style.zIndex='-9999';}"
        )

    def _clicar_js(
        self,
        elemento,
        contexto: str = "",
        scroll_antes: bool = True,
        esconder_rodape: bool = True,
        max_tentativas: int = 3,
    ) -> bool:
        """
        Clique via JavaScript — evita ElementClickInterceptedException.
        Usa scroll, esconde rodapé e retry. Nunca usa elemento.click() do Selenium.
        """
        for tentativa in range(1, max_tentativas + 1):
            try:
                if scroll_antes:
                    self.driver.execute_script(
                        "arguments[0].scrollIntoView({block:'center',inline:'center',behavior:'instant'});",
                        elemento,
                    )
                    time.sleep(0.25)
                if esconder_rodape:
                    self._esconder_rodape_fixo()
                    time.sleep(0.15)
                self.driver.execute_script("arguments[0].click();", elemento)
                return True
            except StaleElementReferenceException:
                if contexto and tentativa < max_tentativas:
                    self.logger.log(
                        f"Elemento obsoleto ao clicar ({contexto}) — tentativa {tentativa}/{max_tentativas}",
                        "WARNING",
                    )
                raise
            except Exception as e:
                if tentativa < max_tentativas and contexto:
                    self.logger.log(
                        f"Clique falhou ({contexto}) — tentativa {tentativa}/{max_tentativas}: {e}",
                        "WARNING",
                    )
                    time.sleep(0.4)
                else:
                    raise
        return False

    # ── Driver ──────────────────────────────────────────────

    def criar_driver(self) -> bool:
        try:
            opts = webdriver.ChromeOptions()

            # Modo anônimo — evita detecção de bot e histórico de sessões anteriores
            opts.add_argument("--incognito")

            # Flags anti-detecção de automação
            opts.add_argument("--disable-blink-features=AutomationControlled")
            opts.add_experimental_option("excludeSwitches", ["enable-automation"])
            opts.add_experimental_option("useAutomationExtension", False)

            opts.add_argument("--no-sandbox")
            opts.add_argument("--disable-dev-shm-usage")
            opts.add_argument("--disable-gpu")
            opts.add_argument("--window-size=1280,900")

            # Selenium Manager (embutido no Selenium 4.6+) detecta a versão do Chrome
            # e baixa o ChromeDriver correto automaticamente — sem dependência externa.
            self.driver = webdriver.Chrome(options=opts)

            # Remove a propriedade navigator.webdriver que delata automação
            self.driver.execute_cdp_cmd(
                "Page.addScriptToEvaluateOnNewDocument",
                {"source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"}
            )

            self.wait = WebDriverWait(self.driver, 60)
            self.logger.log("Navegador iniciado em modo anônimo (anti-bot).", "SUCCESS")
            return True
        except Exception as exc:
            self.logger.log(f"Erro ao iniciar navegador: {exc}", "ERROR")
            return False

    def fechar(self):
        try:
            if self.driver:
                self.driver.quit()
                self.driver = None
        except Exception:
            pass

    # ── Controle de fluxo ───────────────────────────────────

    def _aguardar_pausa(self):
        while self.pausado and not self.parar:
            time.sleep(0.3)

    # ── Login ────────────────────────────────────────────────

    def fazer_login(self, usuario: str, senha: str) -> bool:
        try:
            self.logger.log("Acessando página de login...", "INFO")
            self.driver.get(URL_LOGIN)
            time.sleep(2)

            campo_u = campo_s = btn = None

            try:
                campo_u = self.wait.until(EC.presence_of_element_located((By.ID, ID_LOGIN_USUARIO)))
                campo_s = self.driver.find_element(By.ID, ID_LOGIN_SENHA)
                btn     = self.driver.find_element(By.ID, ID_LOGIN_BTN)
            except Exception:
                for inp in self.driver.find_elements(By.TAG_NAME, "input"):
                    id_  = (inp.get_attribute("id")    or "").lower()
                    typ_ = (inp.get_attribute("type")  or "").lower()
                    val_ = (inp.get_attribute("value") or "").lower()
                    if "usuario" in id_ or "user" in id_:
                        campo_u = inp
                    elif "senha" in id_ or "password" in id_ or typ_ == "password":
                        campo_s = inp
                    elif typ_ in ("submit", "button") or "entrar" in val_ or "ok" in val_:
                        btn = inp

            if not all([campo_u, campo_s, btn]):
                raise Exception("Campos de login não encontrados.")

            self.logger.log("Preenchendo credenciais...", "INFO")
            campo_u.clear(); campo_u.send_keys(usuario)
            campo_s.clear(); campo_s.send_keys(senha)
            self._clicar_js(btn, contexto="Login", scroll_antes=True, esconder_rodape=True)
            time.sleep(3)

            self.logger.log("Login realizado com sucesso.", "SUCCESS")
            return True
        except Exception as exc:
            self.logger.log(f"Erro no login: {exc}", "ERROR")
            return False

    # ── Navegação especial para o sistema ───────────────────

    def navegar_sistema(self) -> bool:
        """
        Fluxo especial exigido pelo sistema:
          1. Navegar para a URL principal
          2. Abrir nova guia com a mesma URL
          3. Fechar a guia original
          4. Continuar na nova guia
        """
        try:
            self.logger.log("Navegando para o sistema de lotes...", "INFO")
            self.driver.get(URL_LOTES)
            time.sleep(3)

            self.logger.log("Abrindo nova guia (fluxo especial do sistema)...", "INFO")
            self.driver.execute_script(f"window.open('{URL_LOTES}', '_blank');")
            time.sleep(2)

            janelas = self.driver.window_handles
            if len(janelas) >= 2:
                self.driver.switch_to.window(janelas[0])
                self.driver.close()
                time.sleep(1)
                self.driver.switch_to.window(self.driver.window_handles[0])

            time.sleep(3)
            self.logger.log("Sistema carregado com sucesso.", "SUCCESS")
            return True
        except Exception as exc:
            self.logger.log(f"Erro ao navegar para o sistema: {exc}", "ERROR")
            return False

    # ── Filtros ──────────────────────────────────────────────

    def aplicar_filtros(
        self,
        data_ini: str,
        data_fim: str,
        tipo_valor: str,
        forma_valor: str = "",
        tipo_fisc_valor: str = "",
        subtipo_valor: str = "",
    ) -> bool:
        try:
            self.logger.log(
                f"Aplicando filtros: {data_ini} até {data_fim} | tipo notif.={tipo_valor} | "
                f"forma={forma_valor or '—'} | tipo fisc.={tipo_fisc_valor or '—'} | subtipo={subtipo_valor or '—'}",
                "INFO"
            )

            campo_ini = self.wait.until(EC.presence_of_element_located((By.ID, ID_DATA_INICIAL)))
            self.driver.execute_script("arguments[0].value = arguments[1];", campo_ini, data_ini)

            campo_fim = self.driver.find_element(By.ID, ID_DATA_FINAL)
            self.driver.execute_script("arguments[0].value = arguments[1];", campo_fim, data_fim)

            # Tipo de notificação — aguardar possível postback após seleção
            select_tipo = Select(self.driver.find_element(By.ID, ID_TIPO_PUBLICACAO))
            select_tipo.select_by_value(tipo_valor)
            self.logger.log("Tipo de publicação selecionado — aguardando atualização da página...", "INFO")
            time.sleep(2 * self.fator_delay)
            self._aguardar_pagina_pronta(timeout=20, contexto="filtros após tipo publicação")

            # Forma de Fiscalização (cascata: altera opções de Tipo Fiscalização)
            if forma_valor != "":
                select_forma = Select(self.driver.find_element(By.ID, ID_FORMA_FISCALIZACAO))
                select_forma.select_by_value(forma_valor)
                self.logger.log("Forma de fiscalização selecionada — aguardando postback...", "INFO")
                time.sleep(2 * self.fator_delay)
                self._aguardar_pagina_pronta(timeout=20, contexto="filtros após forma")

            # Tipo Fiscalização (cascata: altera opções de Subtipo)
            if tipo_fisc_valor != "":
                select_tipo_fisc = Select(self.driver.find_element(By.ID, ID_TIPO_FISCALIZACAO))
                select_tipo_fisc.select_by_value(tipo_fisc_valor)
                self.logger.log("Tipo de fiscalização selecionado — aguardando postback...", "INFO")
                time.sleep(2 * self.fator_delay)
                self._aguardar_pagina_pronta(timeout=20, contexto="filtros após tipo fiscalização")

            # Subtipo Fiscalização
            if subtipo_valor != "":
                select_subtipo = Select(self.driver.find_element(By.ID, ID_SUBTIPO_FISCALIZACAO))
                select_subtipo.select_by_value(subtipo_valor)
                self.logger.log("Subtipo de fiscalização selecionado.", "INFO")
                time.sleep(0.5 * self.fator_delay)

            self._aguardar_pagina_pronta(timeout=20, contexto="filtros pré-pesquisar")
            btn_pesq = self.wait.until(EC.presence_of_element_located((By.ID, ID_BTN_PESQUISAR)))
            self._clicar_js(btn_pesq, contexto="Pesquisar", scroll_antes=True, esconder_rodape=True)
            self.logger.log("Pesquisar clicado — aguardando tabela de lotes...", "INFO")
            self._aguardar_pagina_pronta(timeout=20, contexto="filtros pós-pesquisar")

            try:
                WebDriverWait(self.driver, 20, poll_frequency=0.5).until(
                    lambda d: d.find_elements(By.CSS_SELECTOR, CSS_BTN_ASSINAR)
                )
                total = len(self.driver.find_elements(By.CSS_SELECTOR, CSS_BTN_ASSINAR))
                self.logger.log(f"Tabela carregada — {total} lote(s) disponível(is).", "SUCCESS")
                return True
            except TimeoutException:
                self.logger.log("Nenhum lote encontrado para os filtros informados.", "WARNING")
                return False

        except Exception as exc:
            self.logger.log(f"Erro ao aplicar filtros: {exc}", "ERROR")
            return False

    # ── Leitura da tabela ────────────────────────────────────

    def ler_lotes_tabela(self) -> list[dict]:
        lotes = []
        try:
            tabela = self.driver.find_element(By.ID, ID_TABELA_LOTES)
            # CSS_SELECTOR mais estável que XPath — evita quebra se estrutura mudar
            linhas = tabela.find_elements(By.CSS_SELECTOR, "tbody tr")
            for i, linha in enumerate(linhas):
                try:
                    cols = linha.find_elements(By.CSS_SELECTOR, "td")
                    if not cols:
                        continue
                    numero_lote = (cols[0].text.strip() if len(cols) > 0 else "").strip()
                    # Ignorar linha de "Nenhum registro encontrado." e linhas sem número de lote válido
                    if not numero_lote:
                        continue
                    if numero_lote.lower() == "nenhum registro encontrado.":
                        continue
                    # Só considerar linhas cuja primeira coluna pareça número de lote (evita rodapé/outras mensagens)
                    if not numero_lote.replace(".", "").replace(",", "").isdigit():
                        continue
                    lotes.append({
                        "index":                i,
                        "numero_lote":          numero_lote,
                        "data_envio":           cols[1].text.strip() if len(cols) > 1 else "—",
                        "tipo_fiscalizacao":    cols[2].text.strip() if len(cols) > 2 else "—",
                        "subtipo_fiscalizacao": cols[3].text.strip() if len(cols) > 3 else "—",
                    })
                except StaleElementReferenceException:
                    continue
                except Exception:
                    continue
            self.logger.log(f"{len(lotes)} lote(s) lido(s) da tabela.", "INFO")
        except Exception as exc:
            self.logger.log(f"Erro ao ler tabela: {exc}", "ERROR")
        return lotes

    # ── Assinatura de UM lote ────────────────────────────────

    def assinar_lote(self, senha: str, numero_lote: str, idx_botao: int = 0) -> bool:
        """
        Clica no botão de assinatura pelo índice real do lote na tabela.
        Após a assinatura o botão fica desabilitado (não some), por isso cada
        lote mantém seu índice original e processamos na ordem 0, 1, 2, ...

        Fluxo:
          1. Clicar no botão Assinar
          2. Aguardar nova guia abrir
          3. Aguardar página de assinatura carregar
          4. Preencher senha
          5. Clicar em Salvar
          6. Aguardar sistema processar (delay maior aqui)
          7. Aguardar MessageBox_ButtonOk aparecer
          8. Clicar em OK
          9. Aguardar guia fechar automaticamente
         10. Voltar para aba principal
        """
        try:
            self._aguardar_pausa()
            if self.parar:
                return False

            # Flag: só vira True depois que a barra de progresso sumiu.
            # Se um crash ocorrer ANTES disso, o lote NÃO foi assinado → erro real.
            # Se ocorrer DEPOIS, o sistema processou → trata como sucesso.
            processamento_concluido = False

            janelas_antes = set(self.driver.window_handles)
            aba_principal = self.driver.current_window_handle

            # ── Etapa 1: clicar no botão Assinar pelo índice real da tabela ──
            # Verificar overlays ANTES de clicar — o modal do lote anterior pode ainda estar visível
            self._aguardar_pagina_pronta(timeout=30, contexto=f"Lote {numero_lote} pré-assinar")

            btn_id = f"{_CP}gdvLotePublicacao_btnAssinarAuto_{idx_botao}"
            self.logger.log(f"Lote {numero_lote}: localizando botão Assinar (índice {idx_botao})...", "INFO")

            clicou = False
            for tentativa in range(1, 4):
                btn = self.wait.until(EC.presence_of_element_located((By.ID, btn_id)))
                try:
                    self._clicar_js(
                        btn,
                        contexto=f"Lote {numero_lote} Assinar",
                        scroll_antes=True,
                        esconder_rodape=True,
                        max_tentativas=1,
                    )
                    clicou = True
                    break
                except Exception as e:
                    self.logger.log(
                        f"Lote {numero_lote}: tentativa {tentativa}/3 de clique falhou — {e}", "WARNING"
                    )
                    if tentativa < 3:
                        time.sleep(0.5)

            if not clicou:
                raise Exception(
                    "Não foi possível clicar no botão Assinar após 3 tentativas — "
                    "rodapé fixo pode estar bloqueando."
                )
            self.logger.log(f"Lote {numero_lote}: botão Assinar clicado.", "INFO")

            # ── Etapa 2: aguardar nova guia de assinatura abrir ──
            # O sistema abre uma nova guia ao clicar no checkmark (✓) da coluna "Assinar Lote".
            # Essa nova guia contém o formulário de senha do certificado.
            self._sleep("apos_clicar_assinar")
            self.logger.log(
                f"Lote {numero_lote}: aguardando nova guia de assinatura abrir (máx 60s)...", "INFO"
            )
            try:
                WebDriverWait(self.driver, 60, poll_frequency=0.5).until(
                    lambda d: len(d.window_handles) > len(janelas_antes)
                )
            except TimeoutException:
                self.logger.log(
                    f"Lote {numero_lote}: [ERRO] a nova guia com o formulário de assinatura "
                    f"não abriu em 60s após clicar no botão ✓ (Assinar Lote) da linha {idx_botao}.",
                    "ERROR"
                )
                return False

            nova_janela = (set(self.driver.window_handles) - janelas_antes).pop()
            self.driver.switch_to.window(nova_janela)
            self.logger.log(f"Lote {numero_lote}: nova guia aberta — aguardando carregar...", "INFO")

            # ── Etapa 3: aguardar página de assinatura carregar ──
            self._sleep("carregar_guia_assinatura")

            # Verificar se Progress_UpdateProgress está visível e aguardar sumir.
            # wings_process_presentation_dashboard_bar é rodapé FIXO e permanente —
            # NÃO é overlay de carregamento e nunca deve ser aguardado.
            self._aguardar_pagina_pronta(timeout=60, contexto=f"Lote {numero_lote}")

            # ── Etapa 4: preencher senha ──
            self.logger.log(f"Lote {numero_lote}: preenchendo senha do certificado...", "INFO")

            campo_senha = WebDriverWait(self.driver, 30, poll_frequency=0.3).until(
                EC.presence_of_element_located((By.ID, ID_SENHA_CERT))
            )

            # Rolar para o centro para garantir visibilidade (evita rodapé fixo)
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", campo_senha)
            time.sleep(0.5)

            # Verificar overlay novamente após o scroll
            self._aguardar_pagina_pronta(timeout=15, contexto=f"Lote {numero_lote} pré-senha")

            # JS focus evita completamente ElementClickInterceptedException
            self.driver.execute_script("arguments[0].focus();", campo_senha)
            time.sleep(0.3)
            campo_senha.clear()
            campo_senha.send_keys(senha)
            self.logger.log(f"Lote {numero_lote}: senha preenchida.", "INFO")

            # ── Etapa 5: pausa antes de clicar Salvar ──
            self._sleep("apos_preencher_senha")

            # ── Etapa 6: clicar em Salvar / Confirmar ──
            # Verificar overlay antes de clicar Salvar
            self._aguardar_pagina_pronta(timeout=15, contexto=f"Lote {numero_lote} pré-salvar")

            self.logger.log(f"Lote {numero_lote}: clicando em Salvar...", "INFO")
            btn_salvar = WebDriverWait(self.driver, 30, poll_frequency=0.3).until(
                EC.presence_of_element_located((By.ID, ID_BTN_SALVAR))
            )
            self._clicar_js(
                btn_salvar,
                contexto=f"Lote {numero_lote} Salvar",
                scroll_antes=True,
                esconder_rodape=True,
            )

            # ── Etapa 7a: aguardar barra de progresso APARECER ──
            # Confirma que o sistema começou a processar.
            # XPath: //*[@id="Progress_UpdateProgress"]/div[2]/div/div/div/div/div/div/span
            ID_PROGRESS = "Progress_UpdateProgress"
            timeout_ok  = DELAYS["aguardar_messagebox"]
            log_cada    = DELAYS["log_aguardando_ok_cada"]

            self.logger.log(
                f"Lote {numero_lote}: Salvar clicado — aguardando barra de progresso aparecer...", "INFO"
            )
            try:
                WebDriverWait(self.driver, 15, poll_frequency=0.3).until(
                    EC.visibility_of_element_located((By.ID, ID_PROGRESS))
                )
                self.logger.log(f"Lote {numero_lote}: barra de progresso iniciada — sistema processando...", "INFO")
            except TimeoutException:
                # Pode não aparecer se o sistema for muito rápido — continua normalmente
                self.logger.log(
                    f"Lote {numero_lote}: barra de progresso não detectada (sistema pode ter respondido rápido).", "INFO"
                )

            # ── Etapa 7b: aguardar barra de progresso SUMIR ──
            # Quando a barra some, o processamento terminou e o botão OK está prestes a aparecer.
            # NÃO É ERRO ficar aguardando aqui — pode durar até 600s.
            self.logger.log(
                f"Lote {numero_lote}: aguardando processamento terminar "
                f"(barra de progresso sumir — timeout máximo: {timeout_ok}s)...", "INFO"
            )
            inicio_espera = time.time()
            ultimo_log    = inicio_espera

            while time.time() - inicio_espera < timeout_ok:
                if self.parar:
                    return False
                try:
                    el = self.driver.find_element(By.ID, ID_PROGRESS)
                    visivel = el.is_displayed()
                except NoSuchElementException:
                    visivel = False

                if not visivel:
                    elapsed = int(time.time() - inicio_espera)
                    self.logger.log(
                        f"Lote {numero_lote}: barra de progresso sumiu ({elapsed}s) — processamento concluído!", "INFO"
                    )
                    processamento_concluido = True  # sistema terminou — a partir daqui crash = sucesso
                    break

                # Log periódico a cada N segundos enquanto aguarda
                agora = time.time()
                if agora - ultimo_log >= log_cada:
                    elapsed = int(agora - inicio_espera)
                    self.logger.log(
                        f"Lote {numero_lote}: processando... ({elapsed}s / {timeout_ok}s)", "INFO"
                    )
                    ultimo_log = agora
                time.sleep(0.5)
            else:
                self.logger.log(
                    f"Lote {numero_lote}: [ERRO] timeout de {timeout_ok}s esgotado — "
                    "a barra de progresso não sumiu. Possível causa: senha incorreta ou sistema travado.",
                    "ERROR"
                )
                return False

            # ── Etapa 7c: aguardar botão OK aparecer e clicar (opcional) ──
            # Pausa extra após progresso sumir — dá tempo para o sistema renderizar o botão OK
            self._sleep("apos_progresso_sumir")

            # Verifica se a guia ainda existe antes de tentar interagir
            if nova_janela not in self.driver.window_handles:
                self.logger.log(
                    f"Lote {numero_lote}: guia fechou sozinha após processamento — considerado sucesso.",
                    "SUCCESS"
                )
            else:
                self.logger.log(f"Lote {numero_lote}: aguardando botão OK aparecer...", "INFO")
                btn_ok = None
                try:
                    btn_ok = WebDriverWait(self.driver, 15, poll_frequency=0.3).until(
                        EC.element_to_be_clickable((By.ID, ID_MESSAGEBOX_OK))
                    )
                except TimeoutException:
                    # OK não apareceu — não é erro, o sistema pode ter fechado a guia ou
                    # processado sem exibir confirmação. Fecha a guia e segue.
                    self.logger.log(
                        f"Lote {numero_lote}: botão OK não apareceu — fechando guia e passando para o próximo.",
                        "WARNING"
                    )
                except Exception:
                    # Guia fechou enquanto aguardávamos — trata como sucesso silencioso
                    self.logger.log(
                        f"Lote {numero_lote}: guia fechou durante espera do OK — seguindo.",
                        "WARNING"
                    )

                if btn_ok is not None:
                    self.logger.log(
                        f"Lote {numero_lote}: botão OK detectado "
                        f"({int(time.time() - inicio_espera)}s de espera total)!", "SUCCESS"
                    )
                    try:
                        self._clicar_js(
                            btn_ok,
                            contexto=f"Lote {numero_lote} OK",
                            scroll_antes=True,
                            esconder_rodape=True,
                        )
                        self.logger.log(f"Lote {numero_lote}: OK clicado.", "INFO")
                    except Exception:
                        self.logger.log(
                            f"Lote {numero_lote}: guia fechou antes de clicar OK — seguindo.", "WARNING"
                        )

                # ── Etapa 9: aguardar guia fechar automaticamente ──
                self._sleep("apos_clicar_ok")
                if nova_janela in self.driver.window_handles:
                    try:
                        self.driver.switch_to.window(nova_janela)
                        WebDriverWait(self.driver, 10, poll_frequency=0.5).until(
                            lambda d: nova_janela not in d.window_handles
                        )
                    except Exception:
                        # Guia pode ter fechado durante o wait — tenta fechar manualmente
                        try:
                            if nova_janela in self.driver.window_handles:
                                self.driver.switch_to.window(nova_janela)
                                self.driver.close()
                                self.logger.log(
                                    f"Lote {numero_lote}: guia fechada manualmente.", "WARNING"
                                )
                        except Exception:
                            pass

            # ── Etapa 10: voltar para aba principal ──
            try:
                if aba_principal in self.driver.window_handles:
                    self.driver.switch_to.window(aba_principal)
                elif self.driver.window_handles:
                    self.driver.switch_to.window(self.driver.window_handles[0])
            except Exception:
                pass

            self._sleep("apos_voltar_aba")
            self.logger.log(f"Lote {numero_lote}: processado com sucesso!", "SUCCESS")
            return True

        except Exception as exc:
            msg_erro = str(exc).strip()
            # Identifica se foi o ChromeDriver perdendo a janela (crash sem mensagem real)
            crash_de_janela = (
                not msg_erro                                           # backtrace vazio (janela fechou)
                or "no such window" in msg_erro.lower()
                or "invalid session id" in msg_erro.lower()
                or "target window already closed" in msg_erro.lower()
                or "no such execution context" in msg_erro.lower()
            )

            if crash_de_janela and processamento_concluido:
                # Sistema processou (barra sumiu) → a guia fechou sozinha depois → é sucesso
                self.logger.log(
                    f"Lote {numero_lote}: guia fechou após processamento — tratado como sucesso.",
                    "WARNING"
                )
                sucesso_final = True
            elif crash_de_janela and not processamento_concluido:
                # Crash ANTES do processamento terminar → lote NÃO foi assinado → erro real
                self.logger.log(
                    f"Lote {numero_lote}: [ERRO] guia fechou ANTES do processamento concluir — lote não assinado.",
                    "ERROR"
                )
                sucesso_final = False
            else:
                tipo = type(exc).__name__
                msg = str(exc).strip() or "(sem mensagem)"
                self.logger.log(
                    f"Lote {numero_lote}: [ERRO] ao assinar — {tipo}: {msg}", "ERROR"
                )
                sucesso_final = False

            try:
                if self.driver and self.driver.window_handles:
                    self.driver.switch_to.window(self.driver.window_handles[0])
            except Exception:
                pass
            return sucesso_final

    # ── Paginação ─────────────────────────────────────────────

    def _ir_proxima_pagina(self) -> bool:
        """
        Tenta navegar para a próxima página da tabela de lotes.
        Retorna True se conseguiu avançar, False se já está na última página.
        """
        try:
            btn = self.driver.find_element(By.ID, ID_BTN_PROX_PAG)
            if not btn.is_displayed() or not btn.is_enabled():
                return False
            self._aguardar_pagina_pronta(timeout=15, contexto="pré-próxima página")
            self._clicar_js(btn, contexto="Próxima página", scroll_antes=True, esconder_rodape=True)
            self._sleep("carregar_proxima_pagina")
            self._aguardar_pagina_pronta(timeout=30, contexto="carregando próxima página")
            return True
        except NoSuchElementException:
            return False
        except Exception as exc:
            self.logger.log(f"Erro ao tentar avançar página: {exc}", "WARNING")
            return False

    def coletar_numeros_lotes_todas_paginas(self) -> set[str]:
        """
        Percorre todas as páginas da tabela e retorna o conjunto de números de lote
        presentes. Deixa o driver na última página ao terminar.
        """
        numeros = set()
        while not self.parar:
            self._aguardar_pagina_pronta(timeout=30, contexto="coletar números")
            lotes = self.ler_lotes_tabela()
            for l in lotes:
                numeros.add(l["numero_lote"])
            if not self._ir_proxima_pagina():
                break
        return numeros

    # ── Loop principal — processa UM lote por vez ────────────

    def processar_lotes(
        self,
        senha: str,
        progress_cb=None,
        stats_cb=None,
        *,
        data_ini: str = "",
        data_fim: str = "",
        tipo_valor: str = "",
        forma_valor: str = "",
        tipo_fisc_valor: str = "",
        subtipo_valor: str = "",
    ) -> tuple[int, int, set[str], dict[str, dict]]:
        self.resultados.clear()
        assinados  = 0
        erros      = 0
        num_pagina = 1
        lote_global = 0  # contador global de lotes processados (todas as páginas)
        set_numeros_inicio: set[str] = set()
        dict_lotes_inicio: dict[str, dict] = {}
        pode_voltar_pagina_1 = bool(data_ini and data_fim and tipo_valor)

        self.logger.log(
            f"{'='*50}\n"
            f"  Iniciando processamento (todas as páginas)\n"
            f"  Fator de delay: {self.fator_delay}x\n"
            f"{'='*50}", "INFO"
        )

        while not self.parar:
            if self._verificar_sessao_expirada():
                self.mensagem_final = (
                    "Sessão expirada. Você foi redirecionado para a tela de login. "
                    "Execute a automação novamente."
                )
                self.parar = True
                self.logger.log(self.mensagem_final, "ERROR")
                break
            # ── Lê os lotes da página atual ──
            self._aguardar_pagina_pronta(timeout=30, contexto=f"Página {num_pagina}")
            lotes = self.ler_lotes_tabela()

            for l in lotes:
                set_numeros_inicio.add(l["numero_lote"])
                dict_lotes_inicio[l["numero_lote"]] = {
                    "data_envio": l.get("data_envio", "—"),
                    "tipo_fiscalizacao": l.get("tipo_fiscalizacao", "—"),
                    "subtipo_fiscalizacao": l.get("subtipo_fiscalizacao", "—"),
                }

            if not lotes:
                if num_pagina == 1:
                    # Todos os lotes foram assinados: página 1 mostra "Nenhum registro encontrado."
                    # Não finalizar aqui — retornar para _executar fechar o navegador, reabrir no
                    # próximo ciclo, reaplicar filtros e rodar a verificação de "lotes que sumiram".
                    self.logger.log(
                        f"\nPágina 1 sem lotes (Nenhum registro encontrado.) — todos assinados neste ciclo. "
                        f"Fechando navegador e reabrindo para verificar lotes que sumiram...",
                        "INFO"
                    )
                    break
                # Página 2+ vazia: lotes "subiram" para páginas anteriores — voltar para página 1
                if pode_voltar_pagina_1:
                    self.logger.log(
                        f"\nPágina {num_pagina} sem lotes — voltando para a página 1 para continuar (lotes podem ter sido realocados)...",
                        "INFO"
                    )
                    if not self.aplicar_filtros(data_ini, data_fim, tipo_valor, forma_valor, tipo_fisc_valor, subtipo_valor):
                        self.logger.log("Não foi possível reaplicar filtros para voltar à página 1.", "WARNING")
                        break
                    num_pagina = 1
                    continue
                self.logger.log(
                    f"\nPágina {num_pagina}: nenhum lote encontrado — processamento finalizado.", "INFO"
                )
                break

            self.logger.log(
                f"\n{'─'*50}\n"
                f"  PÁGINA {num_pagina}  —  {len(lotes)} lote(s) encontrado(s)\n"
                f"{'─'*50}", "INFO"
            )

            for i, lote in enumerate(lotes):
                self._aguardar_pausa()
                if self.parar:
                    self.logger.log("Processamento interrompido pelo usuário.", "WARNING")
                    break
                if self._verificar_sessao_expirada():
                    self.mensagem_final = (
                        "Sessão expirada. Você foi redirecionado para a tela de login. "
                        "Execute a automação novamente."
                    )
                    self.parar = True
                    self.logger.log(self.mensagem_final, "ERROR")
                    break

                lote_global += 1
                numero = lote["numero_lote"]

                if progress_cb:
                    progress_cb(
                        f"Página {num_pagina} — lote {i + 1}/{len(lotes)}  (Nº {numero})"
                    )

                self.logger.log(
                    f"\n--- Pg.{num_pagina} | LOTE {i + 1}/{len(lotes)} (global #{lote_global}): Nº {numero} ---",
                    "INFO"
                )

                # Último lote da página: pequeno delay para tabela estabilizar (evita stale element)
                if i == len(lotes) - 1:
                    self.logger.log("Último lote da página — aguardando estabilizar...", "INFO")
                    time.sleep(1.0 * self.fator_delay)

                # Usa o índice real do lote na tabela (botão fica desabilitado após assinar); retry 1x se falhar
                sucesso = self.assinar_lote(senha, numero, idx_botao=lote["index"])
                if not sucesso:
                    self.logger.log(f"Lote {numero}: tentando novamente (2ª tentativa)...", "WARNING")
                    self._fechar_guia_assinatura_orphan()
                    time.sleep(1.0 * self.fator_delay)
                    sucesso = self.assinar_lote(senha, numero, idx_botao=lote["index"])

                hora = datetime.now().strftime("%H:%M:%S")
                resultado = {**lote, "status": "ASSINADO" if sucesso else "ERRO", "hora": hora}
                self.resultados.append(resultado)

                if sucesso:
                    assinados += 1
                else:
                    erros += 1

                if stats_cb:
                    stats_cb(assinados, erros, resultado)

                # Pausa entre lotes (exceto após o último da página)
                if i < len(lotes) - 1 and not self.parar:
                    status_texto = "✔ assinado" if sucesso else "✘ erro"
                    self.logger.log(
                        f"Lote {numero} concluído ({status_texto}) — pausa antes do próximo...",
                        "INFO" if sucesso else "WARNING"
                    )
                    self._sleep("entre_lotes")

            if self.parar:
                break

            # ── Tenta avançar para a próxima página ──
            self.logger.log(f"\nPágina {num_pagina} finalizada — verificando próxima página...", "INFO")
            if self._ir_proxima_pagina():
                num_pagina += 1
                self.logger.log(f"Avançou para a página {num_pagina}.", "INFO")
            else:
                self.logger.log("Última página atingida — processamento finalizado.", "INFO")
                break

        self.logger.log(
            f"\n{'='*50}\n"
            f"  Processamento encerrado\n"
            f"  Páginas percorridas: {num_pagina}\n"
            f"  Total processado:    {lote_global} lote(s)\n"
            f"  Assinados: {assinados}  |  Erros: {erros}\n"
            f"{'='*50}", "INFO"
        )
        return assinados, erros, set_numeros_inicio, dict_lotes_inicio


# ─────────────────────────────────────────────────────────────
# Interface Gráfica
# ─────────────────────────────────────────────────────────────

VERDE    = "#27ae60"
VERMELHO = "#e74c3c"
AZUL     = "#2980b9"
AMARELO  = "#f39c12"
ESCURO   = "#2c3e50"
CINZA_F  = "#f0f0f0"
CINZA_E  = "#ecf0f1"

# Fonte principal (Segoe UI no Windows é mais moderna; em outros SO pode cair para padrão)
FONTE_FAMILY = "Segoe UI"


class InterfaceGrafica:

    # Tamanho fixo da tela de login
    TAMANHO_LOGIN = (480, 380)

    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Assinatura de Lotes DOU — SIFAMA / ANTT")
        self.root.resizable(True, True)
        self.root.configure(bg=CINZA_F)
        self.root.minsize(self.TAMANHO_LOGIN[0], self.TAMANHO_LOGIN[1])
        self.root.geometry(f"{self.TAMANHO_LOGIN[0]}x{self.TAMANHO_LOGIN[1]}")

        self.automacao: AutomacaoAssinaturaLotes | None = None
        self.thread_automacao: threading.Thread | None  = None

        self.usuario_logado = ""
        self.senha_logada   = ""

         # Momento em que o usuário clicou em "Iniciar" — usado para calcular o tempo total
        self.inicio_execucao: float | None = None
        self.timer_total_pausado: float = 0.0
        self.timer_pausado: bool = False
        self.timer_pause_start: float | None = None
        self.timer_rodando: bool = False
        self._timer_after_id: str | None = None

        self.logger = Logger(callback=self._cb_log)

        self._tela_login()
        w, h = self.TAMANHO_LOGIN
        self._centralizar_janela(w, h)

    def _centralizar_janela(self, largura: int, altura: int, offset_y: int = -60):
        """Aplica geometry com largura/altura e centraliza no monitor.
        offset_y desloca verticalmente (negativo = mais para cima)."""
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        x = max(0, (sw - largura) // 2)
        y = max(0, (sh - altura) // 2 + offset_y)
        self.root.geometry(f"{largura}x{altura}+{x}+{y}")

    def _ajustar_tela_principal(self):
        """Calcula o tamanho real necessário para a tela principal e centraliza.
        Chamado via after() para garantir que o layout já foi calculado pelo Tkinter."""
        self.root.update_idletasks()
        req_w = self.root.winfo_reqwidth()
        req_h = self.root.winfo_reqheight()
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        # Adiciona pequena margem e limita ao tamanho útil da tela (desconta barra de tarefas)
        w = min(req_w + 24, sw - 20)
        h = min(req_h + 24, sh - 80)
        self.root.minsize(min(w, 1000), min(h, 700))
        self._centralizar_janela(w, h)

    def _configurar_hover_botao(self, botao: tk.Button, cor_normal: str, cor_hover: str):
        """Efeito hover: clareia o botão ao passar o mouse (só quando habilitado)."""
        def _on_enter(_):
            if botao["state"] == "normal":
                botao.config(bg=cor_hover)
        def _on_leave(_):
            if botao["state"] == "normal":
                botao.config(bg=cor_normal)
        botao.bind("<Enter>", _on_enter)
        botao.bind("<Leave>", _on_leave)

    # ── Tela de login ────────────────────────────────────────

    def _tela_login(self):
        self.frame_login = tk.Frame(self.root, bg=CINZA_F)
        self.frame_login.pack(fill=tk.BOTH, expand=True)

        center = tk.Frame(self.frame_login, bg=CINZA_F)
        center.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

        tk.Label(center, text="Assinatura de Lotes DOU",
                 font=(FONTE_FAMILY, 20, "bold"), bg=CINZA_F, fg=ESCURO).pack(pady=(0, 6))
        tk.Label(center, text="SIFAMA  ·  ANTT",
                 font=(FONTE_FAMILY, 10), bg=CINZA_F, fg="#95a5a6").pack(pady=(0, 32))

        box = tk.Frame(center, bg="white", relief=tk.GROOVE, bd=2, padx=40, pady=32)
        box.pack()

        for row, (label, attr, show) in enumerate([
            ("Usuário:", "entry_usuario", ""),
            ("Senha:",   "entry_senha",   "*"),
        ]):
            tk.Label(box, text=label, bg="white", font=(FONTE_FAMILY, 10), anchor="w"
                     ).grid(row=row, column=0, sticky="w", pady=10, padx=(0, 12))
            entry = tk.Entry(box, width=28, font=(FONTE_FAMILY, 10), show=show)
            entry.grid(row=row, column=1, padx=(0, 0), pady=10)
            setattr(self, attr, entry)

        self.entry_senha.bind("<Return>", lambda _: self._fazer_login())

        btn_entrar = tk.Button(
            box, text="Entrar", command=self._fazer_login,
            bg=AZUL, fg="white", font=(FONTE_FAMILY, 10, "bold"),
            relief=tk.FLAT, cursor="hand2", width=20, pady=10,
        )
        btn_entrar.grid(row=2, column=0, columnspan=2, pady=(24, 0))
        self._configurar_hover_botao(btn_entrar, AZUL, "#3498db")

    # ── Tela principal ───────────────────────────────────────
    

    def _tela_principal(self):
        self.frame_principal = tk.Frame(self.root, bg=CINZA_F)

        # Cabeçalho
        hdr = tk.Frame(self.frame_principal, bg=ESCURO, height=58)
        hdr.pack(fill=tk.X)
        hdr.pack_propagate(False)
        hi = tk.Frame(hdr, bg=ESCURO)
        hi.pack(fill=tk.BOTH, expand=True, padx=20)
        tk.Label(hi, text="Assinatura de Lotes DOU", font=(FONTE_FAMILY, 14, "bold"),
                 bg=ESCURO, fg="white").pack(side=tk.LEFT, pady=16)
        tk.Label(hi, text=f"  |  Usuário: {self.usuario_logado}",
                 font=(FONTE_FAMILY, 9), bg=ESCURO, fg="#bdc3c7").pack(side=tk.LEFT, pady=16)
        btn_sair = tk.Button(hi, text="Sair", command=self._sair,
                             bg=VERMELHO, fg="white", font=(FONTE_FAMILY, 9, "bold"),
                             relief=tk.FLAT, cursor="hand2", padx=16, pady=8)
        btn_sair.pack(side=tk.RIGHT, pady=12)
        self._configurar_hover_botao(btn_sair, VERMELHO, "#c0392b")

        # Conteúdo (padding maior para não colar nas bordas)
        body = tk.Frame(self.frame_principal, bg=CINZA_F)
        body.pack(fill=tk.BOTH, expand=True, padx=20, pady=12)

        # ── Filtros ──────────────────────────────────────────
        lf_filtros = tk.LabelFrame(body, text="  Filtros de Pesquisa  ",
                                  bg=CINZA_F, font=(FONTE_FAMILY, 10, "bold"),
                                  relief=tk.GROOVE, bd=1, padx=14, pady=12)
        lf_filtros.pack(fill=tk.X, pady=(0, 10))

        row_f = tk.Frame(lf_filtros, bg=CINZA_F)
        row_f.pack(fill=tk.X)

        today = datetime.now().strftime("%d/%m/%Y")

        def _on_data_keyrelease(event, entry_widget):
            txt = entry_widget.get()
            fmt = _formatar_data_dd_mm_yyyy(txt)
            if fmt != txt:
                entry_widget.delete(0, tk.END)
                entry_widget.insert(0, fmt)
                entry_widget.icursor(len(fmt))

        for col, (lbl, attr, default) in enumerate([
            ("Data Início:", "entry_data_ini", today),
            ("Data Fim:",    "entry_data_fim", today),
        ]):
            tk.Label(row_f, text=lbl, bg=CINZA_F, font=(FONTE_FAMILY, 9)
                     ).grid(row=0, column=col * 2, sticky="w", padx=(0, 6), pady=4)
            entry = tk.Entry(row_f, width=12, font=(FONTE_FAMILY, 10))
            entry.insert(0, default)
            entry.grid(row=0, column=col * 2 + 1, padx=(0, 22), pady=4)
            entry.bind("<KeyRelease>", lambda e, w=entry: _on_data_keyrelease(e, w))
            setattr(self, attr, entry)

        tk.Label(row_f, text="Tipo de Notificação:", bg=CINZA_F, font=(FONTE_FAMILY, 9)
                 ).grid(row=0, column=4, sticky="w", padx=(0, 6), pady=4)
        self.combo_tipo = ttk.Combobox(
            row_f, values=list(TIPOS_PUBLICACAO.keys()),
            state="readonly", width=26, font=(FONTE_FAMILY, 9)
        )
        self.combo_tipo.set(list(TIPOS_PUBLICACAO.keys())[0])
        self.combo_tipo.grid(row=0, column=5, pady=4)

        # Forma, Tipo e Subtipo de Fiscalização (cascata)
        row_f2 = tk.Frame(lf_filtros, bg=CINZA_F)
        row_f2.pack(fill=tk.X, pady=(12, 0))
        lista_formas_txt = [t for _v, t in LISTA_FORMAS]
        lista_tipos_fisc_txt = [t for _v, t in LISTA_TIPOS_FISC]
        self._subtipo_opcoes_atual: list[tuple[str, str]] = [("", "--Selecione--")]

        tk.Label(row_f2, text="Forma de Fiscalização:", bg=CINZA_F, font=(FONTE_FAMILY, 9)
                 ).grid(row=0, column=0, sticky="w", padx=(0, 6), pady=4)
        self.combo_forma = ttk.Combobox(
            row_f2, values=lista_formas_txt, state="readonly", width=14, font=(FONTE_FAMILY, 9)
        )
        self.combo_forma.set("--Selecione--")
        self.combo_forma.grid(row=0, column=1, padx=(0, 16), pady=4)

        tk.Label(row_f2, text="Tipo Fiscalização:", bg=CINZA_F, font=(FONTE_FAMILY, 9)
                 ).grid(row=0, column=2, sticky="w", padx=(0, 6), pady=4)
        self.combo_tipo_fisc = ttk.Combobox(
            row_f2, values=lista_tipos_fisc_txt, state="readonly", width=24, font=(FONTE_FAMILY, 9)
        )
        self.combo_tipo_fisc.set("--Selecione--")
        self.combo_tipo_fisc.grid(row=0, column=3, padx=(0, 16), pady=4)
        self.combo_tipo_fisc.bind("<<ComboboxSelected>>", self._on_tipo_fisc_change)

        tk.Label(row_f2, text="Subtipo Fiscalização:", bg=CINZA_F, font=(FONTE_FAMILY, 9)
                 ).grid(row=0, column=4, sticky="w", padx=(0, 6), pady=4)
        self.combo_subtipo = ttk.Combobox(
            row_f2, values=["--Selecione--"], state="readonly", width=42, font=(FONTE_FAMILY, 9)
        )
        self.combo_subtipo.set("--Selecione--")
        self.combo_subtipo.grid(row=0, column=5, pady=4)

        # ── Configuração de velocidade ───────────────────────
        lf_vel = tk.LabelFrame(body, text="  Velocidade / Modo de Execução  ",
                               bg=CINZA_F, font=(FONTE_FAMILY, 10, "bold"),
                               relief=tk.GROOVE, bd=1, padx=14, pady=10)
        lf_vel.pack(fill=tk.X, pady=(0, 10))

        vel_row = tk.Frame(lf_vel, bg=CINZA_F)
        vel_row.pack(fill=tk.X)

        tk.Label(vel_row, text="Fator de delay:", bg=CINZA_F,
                 font=(FONTE_FAMILY, 9)).pack(side=tk.LEFT, padx=(0, 8))

        self.spin_delay = tk.Spinbox(
            vel_row, from_=0.5, to=5.0, increment=0.5,
            width=5, font=(FONTE_FAMILY, 10), format="%.1f",
        )
        self.spin_delay.delete(0, tk.END)
        self.spin_delay.insert(0, "2.0")
        self.spin_delay.pack(side=tk.LEFT)

        tk.Label(vel_row,
                 text="  ×  (0.5 = rápido · 1.0 = normal · 2.0 = teste · 3.0+ = lento)",
                 bg=CINZA_F, font=(FONTE_FAMILY, 8), fg="#7f8c8d").pack(side=tk.LEFT, padx=(6, 0))

        vel_row2 = tk.Frame(lf_vel, bg=CINZA_F)
        vel_row2.pack(fill=tk.X, pady=(8, 0))

        tk.Label(vel_row2, text="Delay senha/cliques:", bg=CINZA_F,
                 font=(FONTE_FAMILY, 9)).pack(side=tk.LEFT, padx=(0, 8))

        self.spin_senha_cliques = tk.Spinbox(
            vel_row2, from_=0.3, to=1.5, increment=0.1,
            width=5, font=(FONTE_FAMILY, 10), format="%.1f",
        )
        self.spin_senha_cliques.delete(0, tk.END)
        self.spin_senha_cliques.insert(0, "0.7")
        self.spin_senha_cliques.pack(side=tk.LEFT)

        tk.Label(vel_row2,
                 text="  ×  (0.3 = mínimo · 0.7 = rápido · 1.0 = normal) — senha e cliques",
                 bg=CINZA_F, font=(FONTE_FAMILY, 8), fg="#7f8c8d").pack(side=tk.LEFT, padx=(6, 0))

        info = (
            f"Delays base:  "
            f"após Assinar={DELAYS['apos_clicar_assinar']}s  |  "
            f"após senha={DELAYS['apos_preencher_senha']}s  |  "
            f"após OK={DELAYS['apos_clicar_ok']}s  |  "
            f"entre lotes={DELAYS['entre_lotes']}s"
        )
        tk.Label(lf_vel, text=info, bg=CINZA_F, font=(FONTE_FAMILY, 8), fg="#95a5a6",
                 justify=tk.LEFT).pack(anchor="w", pady=(6, 0))

        # ── Estatísticas ─────────────────────────────────────
        frame_stat = tk.Frame(body, bg=CINZA_E, relief=tk.GROOVE, bd=1)
        frame_stat.pack(fill=tk.X, pady=(0, 10))
        si = tk.Frame(frame_stat, bg=CINZA_E)
        si.pack(padx=16, pady=10)

        for lbl, attr, cor in [
            ("Assinados:", "lbl_assinados", VERDE),
            ("Erros:",     "lbl_erros",     VERMELHO),
        ]:
            tk.Label(si, text=lbl, bg=CINZA_E, font=(FONTE_FAMILY, 9)).pack(side=tk.LEFT)
            lbl_val = tk.Label(si, text="0", bg=CINZA_E, font=(FONTE_FAMILY, 11, "bold"), fg=cor)
            lbl_val.pack(side=tk.LEFT, padx=(4, 24))
            setattr(self, attr, lbl_val)

        self.lbl_progresso = tk.Label(si, text="Aguardando início...",
                                      bg=CINZA_E, font=(FONTE_FAMILY, 9), fg=AZUL)
        self.lbl_progresso.pack(side=tk.LEFT, padx=(10, 0))

        tk.Label(si, text="Tempo:", bg=CINZA_E, font=(FONTE_FAMILY, 9)).pack(side=tk.LEFT, padx=(16, 6))
        self.lbl_tempo = tk.Label(si, text="00:00:00", bg=CINZA_E, font=(FONTE_FAMILY, 10, "bold"), fg=ESCURO)
        self.lbl_tempo.pack(side=tk.LEFT)

        # ── Lotes que sumiram ────────────────────────────────
        lf_sumidos = tk.LabelFrame(body, text="  Lotes que sumiram (último ciclo)  ",
                                   bg=CINZA_F, font=(FONTE_FAMILY, 10, "bold"),
                                   relief=tk.GROOVE, bd=1, padx=12, pady=10)
        lf_sumidos.pack(fill=tk.X, pady=(0, 10))
        cols_sumidos = ("Hora", "Nº Lote", "Data Envio", "Tipo Fiscalização", "Subtipo", "Status")
        widths_sumidos = [72, 85, 92, 155, 130, 95]
        self.tree_sumidos = ttk.Treeview(lf_sumidos, columns=cols_sumidos, show="headings", height=4)
        for c, w in zip(cols_sumidos, widths_sumidos):
            self.tree_sumidos.heading(c, text=c)
            self.tree_sumidos.column(c, width=w, anchor="center")
        sb_sumidos = ttk.Scrollbar(lf_sumidos, orient=tk.VERTICAL, command=self.tree_sumidos.yview)
        self.tree_sumidos.configure(yscrollcommand=sb_sumidos.set)
        self.tree_sumidos.pack(side=tk.LEFT, fill=tk.X, expand=True)
        sb_sumidos.pack(side=tk.RIGHT, fill=tk.Y)

        # ── Tabela de auditoria ──────────────────────────────
        lf_audit = tk.LabelFrame(body, text="  Auditoria — Lotes Processados  ",
                                 bg=CINZA_F, font=(FONTE_FAMILY, 10, "bold"),
                                 relief=tk.GROOVE, bd=1, padx=10, pady=8)
        lf_audit.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        cols_tree = ("Hora", "Nº Lote", "Data Envio", "Tipo Fiscalização", "Subtipo", "Status")
        widths    = [  72,      85,         92,              155,              130,       95  ]
        self.tree = ttk.Treeview(lf_audit, columns=cols_tree, show="headings", height=8)
        for c, w in zip(cols_tree, widths):
            self.tree.heading(c, text=c)
            self.tree.column(c, width=w, anchor="center")
        self.tree.tag_configure("ok",  foreground=VERDE)
        self.tree.tag_configure("err", foreground=VERMELHO)
        sb = ttk.Scrollbar(lf_audit, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb.pack(side=tk.RIGHT, fill=tk.Y)

        # ── Logs ─────────────────────────────────────────────
        lf_log = tk.LabelFrame(body, text="  Logs  ",
                               bg=CINZA_F, font=(FONTE_FAMILY, 10, "bold"),
                               relief=tk.GROOVE, bd=1, padx=10, pady=6)
        lf_log.pack(fill=tk.X, pady=(0, 8))
        self.txt_log = tk.Text(lf_log, height=5, font=("Consolas", 9),
                               bg="#1e1e1e", fg="#d4d4d4", state=tk.DISABLED)
        sb_log = ttk.Scrollbar(lf_log, orient=tk.VERTICAL, command=self.txt_log.yview)
        self.txt_log.configure(yscrollcommand=sb_log.set)
        self.txt_log.pack(side=tk.LEFT, fill=tk.X, expand=True)
        sb_log.pack(side=tk.RIGHT, fill=tk.Y)

        # ── Barra de progresso ───────────────────────────────
        self.progress = ttk.Progressbar(body, mode="indeterminate")
        self.progress.pack(fill=tk.X, pady=(0, 12))

        # ── Botões de controle ───────────────────────────────
        frame_btns = tk.Frame(body, bg=CINZA_F)
        frame_btns.pack(fill=tk.X, pady=(4, 0))

        btn_cfg = [
            ("▶  Iniciar",   VERDE,    self._iniciar,    "btn_iniciar",   tk.NORMAL,   "#2ecc71"),
            ("⏸  Pausar",    AMARELO,  self._pausar,     "btn_pausar",    tk.DISABLED, "#f1c40f"),
            ("▶  Continuar", AZUL,     self._continuar,  "btn_continuar", tk.DISABLED, "#3498db"),
            ("⏹  Parar",     VERMELHO, self._parar,      "btn_parar",     tk.DISABLED, "#c0392b"),
        ]
        for txt, cor, cmd, attr, estado, cor_hover in btn_cfg:
            b = tk.Button(frame_btns, text=txt, command=cmd,
                          bg=cor, fg="white", font=(FONTE_FAMILY, 10, "bold"),
                          relief=tk.FLAT, cursor="hand2",
                          padx=20, pady=10, state=estado)
            b.pack(side=tk.LEFT, padx=12)
            setattr(self, attr, b)
            self._configurar_hover_botao(b, cor, cor_hover)

    # ── Callbacks ────────────────────────────────────────────

    def _cb_log(self, linha: str, tipo: str):
        def _do():
            self.txt_log.config(state=tk.NORMAL)
            start = self.txt_log.index(tk.END)
            self.txt_log.insert(tk.END, linha + "\n")
            end = self.txt_log.index(tk.END)
            tag_map = {
                "ERROR":   ("_err",  VERMELHO),
                "SUCCESS": ("_ok",   VERDE),
                "WARNING": ("_warn", AMARELO),
            }
            if tipo in tag_map:
                tag, cor = tag_map[tipo]
                self.txt_log.tag_add(tag, f"{start} linestart", end)
                self.txt_log.tag_config(tag, foreground=cor)
            self.txt_log.config(state=tk.DISABLED)
            self.txt_log.see(tk.END)
        self.root.after(0, _do)

    def _cb_stats(self, assinados: int, erros: int, resultado: dict):
        def _do():
            self.lbl_assinados.config(text=str(assinados))
            self.lbl_erros.config(text=str(erros))
            tag = "ok" if resultado["status"] == "ASSINADO" else "err"
            self.tree.insert("", 0, values=(
                resultado.get("hora", ""),
                resultado.get("numero_lote", ""),
                resultado.get("data_envio", ""),
                resultado.get("tipo_fiscalizacao", ""),
                resultado.get("subtipo_fiscalizacao", ""),
                resultado.get("status", ""),
            ), tags=(tag,))
        self.root.after(0, _do)

    def _cb_progresso(self, msg: str):
        self.root.after(0, lambda: self.lbl_progresso.config(text=msg))

    def _on_tipo_fisc_change(self, event=None):
        """Atualiza as opções do combo Subtipo conforme o Tipo de Fiscalização selecionado."""
        texto_tipo = self.combo_tipo_fisc.get()
        valor_tipo = next((v for v, t in LISTA_TIPOS_FISC if t == texto_tipo), "")
        self._subtipo_opcoes_atual = SUBTIPOS_POR_TIPO.get(valor_tipo, [("", "--Selecione--")])
        self.combo_subtipo["values"] = [t for _v, t in self._subtipo_opcoes_atual]
        self.combo_subtipo.set("--Selecione--")

    def _montar_mensagem_final(
        self,
        ciclo: int,
        total_assinados: int,
        total_erros: int,
        list_total_resultados: list[dict],
    ) -> str:
        """Monta mensagem de conclusão com resumo de erros e exporta XLSX se houver resultados."""
        lista_erros = [r["numero_lote"] for r in list_total_resultados if r.get("status") == "ERRO"]
        if lista_erros:
            self.logger.log(
                f"Lotes com ERRO ({len(lista_erros)}): {', '.join(sorted(lista_erros))}",
                "WARNING"
            )
        caminho_xlsx = ""
        if OPENPYXL_DISPONIVEL and list_total_resultados:
            pasta_base = os.path.join(os.path.dirname(os.path.abspath(__file__)), "resultados")
            os.makedirs(pasta_base, exist_ok=True)
            nome_xlsx = f"resultado_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            caminho_xlsx = os.path.join(pasta_base, nome_xlsx)
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Lotes"
                headers = ("Hora", "Nº Lote", "Data Envio", "Tipo Fiscalização", "Subtipo", "Status")
                for col, h in enumerate(headers, 1):
                    c = ws.cell(row=1, column=col, value=h)
                    c.font = Font(bold=True)
                    c.alignment = Alignment(horizontal="center")
                for row_idx, r in enumerate(list_total_resultados, 2):
                    ws.cell(row=row_idx, column=1, value=r.get("hora", ""))
                    ws.cell(row=row_idx, column=2, value=r.get("numero_lote", ""))
                    ws.cell(row=row_idx, column=3, value=r.get("data_envio", ""))
                    ws.cell(row=row_idx, column=4, value=r.get("tipo_fiscalizacao", ""))
                    ws.cell(row=row_idx, column=5, value=r.get("subtipo_fiscalizacao", ""))
                    ws.cell(row=row_idx, column=6, value=r.get("status", ""))
                wb.save(caminho_xlsx)
                self.logger.log(f"Resultado exportado para: {caminho_xlsx}", "INFO")
            except Exception as ex:
                self.logger.log(f"Não foi possível salvar XLSX: {ex}", "WARNING")
                caminho_xlsx = ""
        msg = (
            f"Processamento concluído!\n\n"
            f"Ciclos executados: {ciclo}\n"
            f"Assinados: {total_assinados}\n"
            f"Erros: {total_erros}"
        )
        if lista_erros:
            msg += f"\n\nLotes com ERRO ({len(lista_erros)}): " + ", ".join(sorted(lista_erros))
        if caminho_xlsx:
            msg += f"\n\nResultado exportado para:\n{caminho_xlsx}"
        return msg

    def _cb_lotes_sumidos(self, ciclo: int, linhas: list[tuple]):
        """Recebe lista de tuplas (hora, numero_lote, data_envio, tipo, subtipo, status) e preenche a tabela."""
        def _do():
            for item in self.tree_sumidos.get_children():
                self.tree_sumidos.delete(item)
            for row in linhas:
                self.tree_sumidos.insert("", tk.END, values=row)
        self.root.after(0, _do)

    # ── Ações ────────────────────────────────────────────────

    def _fazer_login(self):
        u = self.entry_usuario.get().strip()
        s = self.entry_senha.get().strip()
        if not u or not s:
            messagebox.showerror("Erro", "Preencha usuário e senha.")
            return
        self.usuario_logado = u
        self.senha_logada   = s
        self._tela_principal()
        self.frame_login.pack_forget()
        self.frame_principal.pack(fill=tk.BOTH, expand=True)
        # Usa after() para garantir que o layout esteja calculado antes de redimensionar
        self.root.after(10, self._ajustar_tela_principal)
        self.logger.log(f"Sessão iniciada para '{u}'.", "SUCCESS")

    def _sair(self):
        if messagebox.askyesno("Sair", "Deseja encerrar a sessão?"):
            if self.automacao:
                self.automacao.fechar()
            self.frame_principal.pack_forget()
            self.frame_login.pack(fill=tk.BOTH, expand=True)
            w, h = self.TAMANHO_LOGIN
            self.root.minsize(w, h)
            self._centralizar_janela(w, h)

    def _iniciar(self):
        data_ini  = self.entry_data_ini.get().strip()
        data_fim  = self.entry_data_fim.get().strip()
        tipo_nome = self.combo_tipo.get()

        if not data_ini or not data_fim:
            messagebox.showerror("Erro", "Preencha as datas de início e fim.")
            return
        if not tipo_nome:
            messagebox.showerror("Erro", "Selecione o Tipo de Notificação.")
            return

        tipo_valor = TIPOS_PUBLICACAO.get(tipo_nome, "")
        if not tipo_valor:
            messagebox.showerror("Erro", "Tipo de publicação inválido.")
            return

        try:
            fator = float(self.spin_delay.get())
            fator = max(0.5, min(5.0, fator))
        except ValueError:
            fator = 2.0

        try:
            fator_senha_cliques = float(self.spin_senha_cliques.get())
            fator_senha_cliques = max(0.3, min(1.5, fator_senha_cliques))
        except ValueError:
            fator_senha_cliques = 0.7

        # Marca o início da execução e inicia o timer ao vivo
        self.inicio_execucao = time.time()
        self.timer_total_pausado = 0.0
        self.timer_pausado = False
        self.timer_pause_start = None
        self.timer_rodando = True
        self.lbl_tempo.config(text="00:00:00")
        self._atualizar_timer()

        self.btn_iniciar.config(state=tk.DISABLED)
        self.btn_pausar.config(state=tk.NORMAL)
        self.btn_parar.config(state=tk.NORMAL)
        self.progress.start(12)
        self.lbl_assinados.config(text="0")
        self.lbl_erros.config(text="0")
        for item in self.tree_sumidos.get_children():
            self.tree_sumidos.delete(item)

        # Criar arquivo de log com timestamp — salvo em logs/ ao lado do script
        pasta_logs = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs")
        nome_log   = f"automacao_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        self.logger.log_file = os.path.join(pasta_logs, nome_log)
        os.makedirs(pasta_logs, exist_ok=True)
        with open(self.logger.log_file, "w", encoding="utf-8") as f:
            f.write(f"Log de Automação — Assinatura de Lotes DOU\n")
            f.write(f"Iniciado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
            f.write(f"Usuário: {self.usuario_logado}\n")
            f.write(f"Filtro: {data_ini} a {data_fim} | {tipo_nome}\n")
            f.write(f"Fator de delay: {fator}x | Delay senha/cliques: {fator_senha_cliques}x\n")
            f.write("=" * 60 + "\n\n")
        self.logger.log(f"Log salvo em: {self.logger.log_file}", "INFO")

        forma_valor = next((v for v, t in LISTA_FORMAS if t == self.combo_forma.get()), "")
        tipo_fisc_valor = next((v for v, t in LISTA_TIPOS_FISC if t == self.combo_tipo_fisc.get()), "")
        subtipo_valor = next((v for v, t in self._subtipo_opcoes_atual if t == self.combo_subtipo.get()), "")

        self.automacao = AutomacaoAssinaturaLotes(
            self.logger,
            fator_delay=fator,
            fator_senha_cliques=fator_senha_cliques,
        )
        self.thread_automacao = threading.Thread(
            target=self._executar,
            args=(data_ini, data_fim, tipo_valor, forma_valor, tipo_fisc_valor, subtipo_valor),
            daemon=True,
        )
        self.thread_automacao.start()

    def _executar(
        self,
        data_ini: str,
        data_fim: str,
        tipo_valor: str,
        forma_valor: str = "",
        tipo_fisc_valor: str = "",
        subtipo_valor: str = "",
    ):
        try:
            total_assinados = 0
            total_erros = 0
            ciclo = 1
            set_inicio_ciclo_anterior = None
            dict_lotes_inicio_ciclo_anterior = None
            list_total_resultados: list[dict] = []

            while not self.automacao.parar:
                self.automacao.logger.log(
                    f"\n{'#'*50}\n  CICLO {ciclo}  —  iniciando navegador e fluxo completo\n{'#'*50}",
                    "INFO"
                )

                if not self.automacao.criar_driver():
                    self.root.after(0, self._finalizar, False, "Erro ao iniciar o navegador.")
                    return

                if not self.automacao.fazer_login(self.usuario_logado, self.senha_logada):
                    self.automacao.fechar()
                    self.root.after(0, self._finalizar, False, "Erro ao fazer login.")
                    return

                if not self.automacao.navegar_sistema():
                    self.automacao.fechar()
                    self.root.after(0, self._finalizar, False, "Erro ao carregar o sistema.")
                    return

                if not self.automacao.aplicar_filtros(data_ini, data_fim, tipo_valor, forma_valor, tipo_fisc_valor, subtipo_valor):
                    self.automacao.fechar()
                    if ciclo == 1:
                        self.root.after(0, self._finalizar, False,
                                        "Nenhum lote encontrado para os filtros informados.")
                    else:
                        if set_inicio_ciclo_anterior and dict_lotes_inicio_ciclo_anterior is not None:
                            lista_ordenada = sorted(set_inicio_ciclo_anterior)
                            qtd = len(lista_ordenada)
                            hora = datetime.now().strftime("%H:%M:%S")
                            linhas = []
                            for num in lista_ordenada:
                                d = dict_lotes_inicio_ciclo_anterior.get(num, {})
                                linhas.append((
                                    hora,
                                    num,
                                    d.get("data_envio", "—"),
                                    d.get("tipo_fiscalizacao", "—"),
                                    d.get("subtipo_fiscalizacao", "—"),
                                    "Sumiu",
                                ))
                            self.automacao.logger.log(
                                f"Ciclo {ciclo - 1}: {qtd} lote(s) sumiram (tabela vazia) — {', '.join(lista_ordenada)}",
                                "SUCCESS"
                            )
                            self.root.after(0, lambda c=ciclo - 1, l=linhas: self._cb_lotes_sumidos(c, l))
                        self.automacao.logger.log(
                            "Nenhum lote encontrado — tabela vazia. Processamento encerrado.",
                            "INFO"
                        )
                        msg = self._montar_mensagem_final(
                            ciclo, total_assinados, total_erros, list_total_resultados
                        )
                        self.root.after(0, self._finalizar, True, msg)
                    return

                # A partir do ciclo 2: verificar quais lotes do ciclo anterior sumiram
                if set_inicio_ciclo_anterior is not None and dict_lotes_inicio_ciclo_anterior is not None:
                    self.automacao.logger.log(
                        "Verificando lotes que sumiram da tabela (após reabrir página)...",
                        "INFO"
                    )
                    set_depois = self.automacao.coletar_numeros_lotes_todas_paginas()
                    lotes_que_sumiram = set_inicio_ciclo_anterior - set_depois
                    qtd = len(lotes_que_sumiram)
                    lista_ordenada = sorted(lotes_que_sumiram)
                    ciclo_verificado = ciclo - 1
                    hora = datetime.now().strftime("%H:%M:%S")
                    linhas = []
                    for num in lista_ordenada:
                        d = dict_lotes_inicio_ciclo_anterior.get(num, {})
                        linhas.append((
                            hora,
                            num,
                            d.get("data_envio", "—"),
                            d.get("tipo_fiscalizacao", "—"),
                            d.get("subtipo_fiscalizacao", "—"),
                            "Sumiu",
                        ))
                    self.automacao.logger.log(
                        f"Ciclo {ciclo_verificado}: {qtd} lote(s) sumiram da relação — {', '.join(lista_ordenada)}",
                        "SUCCESS" if qtd else "INFO"
                    )
                    self.root.after(0, lambda c=ciclo_verificado, l=linhas: self._cb_lotes_sumidos(c, l))
                    # Voltar para a primeira página para processar
                    if not self.automacao.aplicar_filtros(data_ini, data_fim, tipo_valor, forma_valor, tipo_fisc_valor, subtipo_valor):
                        self.automacao.fechar()
                        self.automacao.logger.log(
                            "Nenhum lote encontrado — tabela vazia. Processamento encerrado.",
                            "INFO"
                        )
                        msg = self._montar_mensagem_final(
                            ciclo, total_assinados, total_erros, list_total_resultados
                        )
                        self.root.after(0, self._finalizar, True, msg)
                        return

                self.automacao.logger.log(
                    f"\n  CICLO {ciclo}  —  processando todos os lotes de todas as páginas\n",
                    "INFO"
                )

                assinados, erros, set_inicio, dict_lotes_inicio = self.automacao.processar_lotes(
                    senha=self.senha_logada,
                    progress_cb=self._cb_progresso,
                    stats_cb=self._cb_stats,
                    data_ini=data_ini,
                    data_fim=data_fim,
                    tipo_valor=tipo_valor,
                    forma_valor=forma_valor,
                    tipo_fisc_valor=tipo_fisc_valor,
                    subtipo_valor=subtipo_valor,
                )
                total_assinados += assinados
                total_erros += erros
                set_inicio_ciclo_anterior = set_inicio
                dict_lotes_inicio_ciclo_anterior = dict_lotes_inicio
                list_total_resultados.extend(list(self.automacao.resultados))

                if self.automacao.parar:
                    self.automacao.fechar()
                    break

                self.automacao.logger.log(
                    f"\nCiclo {ciclo} finalizado — fechando navegador para limpar cache e reabrir...",
                    "INFO"
                )
                self.automacao.fechar()
                self.automacao._sleep("apos_fechar_navegador")

                ciclo += 1

            if getattr(self.automacao, "mensagem_final", None):
                self.root.after(0, self._finalizar, False, self.automacao.mensagem_final)
            elif not self.automacao.parar:
                msg = self._montar_mensagem_final(
                    ciclo, total_assinados, total_erros, list_total_resultados
                )
                self.root.after(0, self._finalizar, True, msg)

        except Exception as exc:
            self.logger.log(f"Erro fatal na automação: {exc}", "ERROR")
            if self.automacao:
                self.automacao.fechar()
            self.root.after(0, self._finalizar, False, f"Erro inesperado:\n{exc}")

    def _atualizar_timer(self):
        if not getattr(self, "timer_rodando", False):
            return
        if self.timer_pausado:
            self._timer_after_id = self.root.after(1000, self._atualizar_timer)
            return
        if self.inicio_execucao is None:
            return
        decorrido = time.time() - self.inicio_execucao - self.timer_total_pausado
        seg = max(0, int(decorrido))
        h, m, s = seg // 3600, (seg % 3600) // 60, seg % 60
        self.lbl_tempo.config(text=f"{h:02d}:{m:02d}:{s:02d}")
        self._timer_after_id = self.root.after(1000, self._atualizar_timer)

    def _finalizar(self, sucesso: bool, mensagem: str = ""):
        self.timer_rodando = False
        if self._timer_after_id is not None:
            self.root.after_cancel(self._timer_after_id)
            self._timer_after_id = None
        self.progress.stop()
        self.btn_iniciar.config(state=tk.NORMAL)
        self.btn_pausar.config(state=tk.DISABLED)
        self.btn_continuar.config(state=tk.DISABLED)
        self.btn_parar.config(state=tk.DISABLED)
        self._cb_progresso("Concluído." if sucesso else "Parado.")

        # Calcula tempo total de execução (apenas tempo em que estava rodando, sem pausa)
        info_tempo = ""
        if self.inicio_execucao is not None:
            total_pausado = self.timer_total_pausado
            if self.timer_pause_start is not None:
                total_pausado += time.time() - self.timer_pause_start
            duracao_seg = max(0, int(time.time() - self.inicio_execucao - total_pausado))
            h = duracao_seg // 3600
            m = (duracao_seg % 3600) // 60
            s = duracao_seg % 60
            if h > 0:
                info_tempo = f"\n\nTempo total: {h:02d}:{m:02d}:{s:02d} (hh:mm:ss)"
            else:
                info_tempo = f"\n\nTempo total: {m:02d}:{s:02d} (mm:ss)"
            self.logger.log(f"Tempo total de execução: {duracao_seg}s.", "INFO")

        if mensagem:
            mensagem_final = mensagem + info_tempo
        else:
            mensagem_final = info_tempo.lstrip("\n") if info_tempo else ""

        if mensagem_final:
            (messagebox.showinfo if sucesso else messagebox.showerror)(
                "Resultado" if sucesso else "Erro", mensagem_final
            )

    def _pausar(self):
        if self.automacao:
            self.automacao.pausado = True
            self.timer_pausado = True
            self.timer_pause_start = time.time()
            self.btn_pausar.config(state=tk.DISABLED)
            self.btn_continuar.config(state=tk.NORMAL)
            self.logger.log("Automação pausada.", "WARNING")

    def _continuar(self):
        if self.automacao:
            self.automacao.pausado = False
            if self.timer_pause_start is not None:
                self.timer_total_pausado += time.time() - self.timer_pause_start
                self.timer_pause_start = None
            self.timer_pausado = False
            self.btn_continuar.config(state=tk.DISABLED)
            self.btn_pausar.config(state=tk.NORMAL)
            self.logger.log("Automação retomada.", "INFO")

    def _parar(self):
        if self.automacao:
            self.automacao.parar   = True
            self.automacao.pausado = False
            self.logger.log("Parando automação...", "WARNING")
            self._finalizar(False, "Automação interrompida pelo usuário.")

    # ── Entry point ──────────────────────────────────────────

    def executar(self):
        self.root.mainloop()


# ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    app = InterfaceGrafica()
    app.executar()
